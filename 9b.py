from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
wb=Workbook()
sheet=wb.active
sheet.title="studentdata"
Name=["RAM", "JANU", "ARYA"]
College=["JIT", "RNSIT", "SJBIT"]
Marks=["85", "10", "101"]
sheet.cell(row=1, coloumn=1).value="name"
sheet.cell(row=1, coloumn=2).value="college"
sheet.cell(row=1, coloumn=3).value="marks"
ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft
        for i in range(2, 5):
            sheet.cell(row=i, coloumn=1).value=Name[i-2]
            sheet.cell(row=i, coloumn=2).value=College[i-2]
            sheet.cell(row=i, coloumn=3).value=Marks[i-2]
            wb.save("9bprogram.xlsx")
            data=pd.read_excel('9bprogram.xlsx')
            print(data)